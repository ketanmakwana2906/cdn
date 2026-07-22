"""
UNIVERSAL chart-data extraction script for TI-vs-wind-speed report charts.

Unlike earlier versions, this makes NO assumptions about image resolution,
plot pixel geometry, or axis min/max -- every one of those is detected
fresh from each image. This was specifically built and validated against
two different chart families with different image sizes AND different
axis ranges (KAD-xx: 1520x847px, y-axis 0-31; MAG-PN-xx: ~1179x605px,
y-axis 0-32), including a chart variant with a shaded "IEC demand
exceeded" overlay region.

PIPELINE PER IMAGE
  1. detect_plot_box       - find the axis border box (contiguity-based,
                              robust to text/gridlines outweighing the
                              border in raw pixel count on some renders)
  2. detect_axis_range     - OCR the 4 corner tick labels (x_min, x_max,
                              y_min, y_max) individually -- far more
                              reliable than OCR'ing the whole label strip
                              at once, which breaks on small/dense renders
  3. calibrate              - fit pixel<->value mapping via least-squares
                              regression across ALL detected gridlines
                              (local-baseline gridline detection, robust to
                              shaded overlay regions), anchored by the
                              OCR'd axis range
  4. detect_legend_colors  - OCR the legend to find the exact "Effective
                              TI" (and "ambient TI") swatch colors fresh
                              per image, not hard-coded
  5. extract + clean curve - color-cluster extraction with red/pink
                              hard-lightness-split disambiguation (same
                              hue, different lightness -- a loose color
                              match lets one bleed into the other at
                              crossing points) + continuity-based outlier
                              rejection
  6. calibration_confidence - validates the fitted fit quality; if it's
                              bad, the image is flagged rather than
                              silently trusted
"""

import cv2
import numpy as np
import pandas as pd
import os
import glob
from scipy.ndimage import median_filter

# ============================================================
# ====================  USER CONFIG  ========================
# ============================================================
# Edit these before running. Everything else in this file can be
# left alone for the chart family this was validated against.

# Full path to the Tesseract OCR executable on THIS machine.
# Leave as None to use whatever is already on your system PATH
# (works out of the box on most Linux/Mac installs where you did
# `apt install tesseract-ocr` or `brew install tesseract`).
#
# Common Windows path (adjust to your actual install location):
#   TESSERACT_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
# Common Mac (Homebrew) path:
#   TESSERACT_CMD = "/opt/homebrew/bin/tesseract"
# Common Linux path:
#   TESSERACT_CMD = "/usr/bin/tesseract"
TESSERACT_CMD = None

# Folder containing the chart images to process.
INPUT_FOLDER = "."

# Which image file extensions to pick up from INPUT_FOLDER.
IMAGE_PATTERNS = ('*.png', '*.jpg', '*.jpeg', '*.PNG', '*.JPG', '*.JPEG')

# The wind speeds (x-axis values) to extract at.
TARGET_X = [3, 4, 5, 6, 7, 8, 9, 9.3, 10, 11.3, 12, 14, 16, 18, 20]

# Name of the output Excel file (written inside INPUT_FOLDER unless
# you give a full path).
OUTPUT_XLSX = "master_output_1.xlsx"

# --- Resolution handling ---
# The pipeline's internal thresholds (gridline-detection deviation,
# regression tolerance, outlier windows) are defined in absolute pixel
# terms and were validated against charts in the ~1150-1550px-wide range.
# Very low-resolution images genuinely lack the pixel detail needed to
# read tick labels and thin curve lines reliably -- below
# MIN_SOURCE_WIDTH_PX, the script refuses rather than silently guessing.
# Images moderately below the validated band are upscaled toward
# CANONICAL_WIDTH_PX first, which was found to meaningfully help; images
# already within/above the validated band are left untouched, since
# resizing an already-good image was found to introduce interpolation
# blur rather than help.
CANONICAL_WIDTH_PX = 1400
MIN_SOURCE_WIDTH_PX = 700
NATIVE_BAND_PX = (1000, 1700)
# ============================================================


def normalize_image(img):
    """Resize toward CANONICAL_WIDTH_PX ONLY when the source is below the
    validated native band. Refuses (returns None) below
    MIN_SOURCE_WIDTH_PX, where even upscaling can't recover detail that
    was never captured in the source image -- genuine information loss,
    not something more processing can fix."""
    h, w = img.shape[:2]
    if w < MIN_SOURCE_WIDTH_PX:
        return None, f"source width {w}px below MIN_SOURCE_WIDTH_PX={MIN_SOURCE_WIDTH_PX}"
    if NATIVE_BAND_PX[0] <= w <= NATIVE_BAND_PX[1]:
        return img, None
    scale = CANONICAL_WIDTH_PX / w
    interp = cv2.INTER_AREA if scale < 1 else cv2.INTER_CUBIC
    return cv2.resize(img, None, fx=scale, fy=scale, interpolation=interp), None

try:
    import pytesseract
    if TESSERACT_CMD:
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
    HAS_OCR = True
except ImportError:
    HAS_OCR = False


# ============================================================
# 1. PLOT BORDER BOX DETECTION
# ============================================================

def _longest_run(bool_arr):
    if not bool_arr.any():
        return 0
    diff = np.diff(np.concatenate(([0], bool_arr.astype(int), [0])))
    starts = np.where(diff == 1)[0]
    ends = np.where(diff == -1)[0]
    return (ends - starts).max() if len(starts) else 0


def detect_plot_box(gray, thresh=150):
    """Find the axis border box using longest-CONTIGUOUS-dark-run per
    row/column, not raw dark-pixel count. Raw count fails on some renders
    where the x-axis tick-label row has more total dark pixels than the
    (thin, but unbroken) border line itself. Threshold=150 (not the
    stricter ~120 used in earlier versions) because border line intensity
    was observed to vary between exactly 0 and 120 depending on the
    render -- 150 keeps a wide safety margin below gridline gray (~215-242)
    while catching both."""
    dark = gray < thresh
    H, W = gray.shape
    row_runs = np.array([_longest_run(dark[r, :]) for r in range(H)])
    col_runs = np.array([_longest_run(dark[:, c]) for c in range(W)])
    row_thresh = 0.5 * row_runs.max()
    col_thresh = 0.5 * col_runs.max()
    rows = np.where(row_runs > row_thresh)[0]
    cols = np.where(col_runs > col_thresh)[0]
    if len(rows) == 0 or len(cols) == 0:
        raise ValueError("Could not detect a plot border box in this image")
    top, bottom = int(rows.min()), int(rows.max())
    left, right = int(cols.min()), int(cols.max())
    return top, bottom, left, right


# ============================================================
# 2. AXIS RANGE DETECTION (OCR, anchored single-label reads)
# ============================================================

def _ocr_single_number(crop, scale=6):
    if not HAS_OCR or crop.size == 0:
        return None
    big = cv2.resize(crop, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
    gray = cv2.cvtColor(big, cv2.COLOR_BGR2GRAY) if big.ndim == 3 else big
    for cfg in ('--psm 6', '--psm 7', '--psm 8'):
        txt = pytesseract.image_to_string(
            gray, config=f'{cfg} -c tessedit_char_whitelist=0123456789').strip()
        if txt.isdigit():
            return int(txt)
    return None


def _ocr_strip_numbers(img, box, axis, margin=40, scale=4):
    """Fallback: OCR the WHOLE tick-label strip (less reliable per-digit
    than a single anchored crop, but useful as a cross-check / backup when
    a single anchor read fails)."""
    top, bottom, left, right = box
    H, W = img.shape[:2]
    if axis == 'x':
        crop = img[bottom + 1:min(bottom + 30, H), max(0, left - margin):min(right + margin, W)]
    else:
        crop = img[max(0, top - margin):bottom + margin, max(0, left - 55):left - 1]
    if crop.size == 0 or not HAS_OCR:
        return []
    big = cv2.resize(crop, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
    gray = cv2.cvtColor(big, cv2.COLOR_BGR2GRAY)
    nums = []
    for cfg in ('--psm 11', '--psm 6'):
        txt = pytesseract.image_to_string(
            gray, config=f'{cfg} -c tessedit_char_whitelist=0123456789.-')
        for tok in txt.replace('\n', ' ').split():
            tok = tok.strip('.-')
            if tok.isdigit() and len(tok) <= 3:
                nums.append(int(tok))
    return nums


def _robust_max_from_list(nums):
    """From a noisy list of OCR'd numbers, take the highest value that is
    corroborated by a nearby value consistent with the ACTUAL tick
    spacing used on this axis -- rejects isolated garbage reads
    (observed: colorized small tick text sometimes OCRs as a wildly
    larger stray number) while correctly handling any tick interval (1,
    2, 5, 10, ...), not just spacing-of-1 axes. A fixed small gap
    tolerance (assuming spacing=1) was found to silently truncate the
    detected max on any axis with wider tick spacing -- e.g. an axis
    labeled 0,5,10,...,100 would incorrectly resolve to max=5, since
    consecutive real labels are 5 apart, exceeding a fixed tolerance."""
    if not nums:
        return None
    vals = sorted(set(nums))
    if len(vals) < 3:
        return vals[-1] if len(vals) >= 1 else None

    # Estimate the true tick spacing from the most common gap between
    # consecutive sorted values (robust to a few missing/garbage reads).
    gaps = np.diff(vals)
    gaps = gaps[gaps > 0]
    if len(gaps) == 0:
        return vals[-1]
    spacing = np.median(gaps)
    tol = max(3, spacing * 1.5)  # generous enough to bridge one missed label

    survivors = [vals[-1]]
    for v in reversed(vals[:-1]):
        if survivors[-1] - v <= tol:
            survivors.append(v)
        elif len(survivors) >= 2:
            break
        else:
            survivors = [v]
    return max(survivors) if len(survivors) >= 2 else vals[-1]


def _count_gridlines(img, box, axis):
    """Probes several candidate rows/columns (not just one) and keeps the
    reading with the most detected gridlines. A single probe can be
    partially corrupted by a shaded overlay region (e.g. 'IEC demand
    exceeded' band), undercounting gridlines and poisoning the axis-range
    arbitration downstream; trying several positions and taking the best
    is robust to any one of them being compromised."""
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY).astype(float)
    top, bottom, left, right = box
    best_n = 0
    if axis == 'x':
        for frac in np.linspace(0.90, 0.995, 8):
            row = top + int((bottom - top) * frac)
            g = _detect_gridlines_local(gray[row, left:right + 1])
            best_n = max(best_n, len(g))
    else:
        for frac in np.linspace(0.005, 0.10, 8):
            col = left + int((right - left) * frac)
            g = _detect_gridlines_local(gray[top:bottom + 1, col])
            best_n = max(best_n, len(g))
    return best_n


def detect_axis_range(img, box, default_x_min=0, default_y_min=0):
    """Detect (x_min, x_max), (y_min, y_max) using TWO independent OCR
    strategies for x_max/y_max:
      (a) whole-strip OCR + outlier-robust max-picking
      (b) a single anchored crop right at the corner

    Both were observed to fail unpredictably in EITHER direction (under-
    read a digit, or over-read one -- e.g. '31' misread as both '4' in
    one case and '39' in another on different images), so neither can be
    trusted alone or blindly combined via max(). Instead, the actual
    number of detected gridlines (a structural, non-OCR signal) is used
    as ground truth to arbitrate: whichever candidate implies a tick
    count closest to the real detected gridline count wins.
    """
    top, bottom, left, right = box
    H, W = img.shape[:2]

    x_max_strip = _robust_max_from_list(_ocr_strip_numbers(img, box, 'x'))
    y_max_strip = _robust_max_from_list(_ocr_strip_numbers(img, box, 'y'))

    x_max_crop = img[bottom + 1:min(bottom + 32, H), max(0, right - 25):min(right + 45, W)]
    y_max_crop = img[max(0, top - 12):top + 12, max(0, left - 40):left - 1]
    x_max_anchor = _ocr_single_number(x_max_crop)
    y_max_anchor = _ocr_single_number(y_max_crop)

    x_min_crop = img[bottom + 1:min(bottom + 32, H), max(0, left - 15):left + 20]
    y_min_crop = img[bottom - 12:bottom + 12, max(0, left - 40):left - 1]
    x_min = _ocr_single_number(x_min_crop)
    y_min = _ocr_single_number(y_min_crop)
    if x_min is None or x_min > 5:
        x_min = default_x_min
    if y_min is None or y_min > 5:
        y_min = default_y_min

    n_x_grid = _count_gridlines(img, box, 'x')
    n_y_grid = _count_gridlines(img, box, 'y')

    def pick_best(candidates, axis_min, n_grid_detected, prefer_index=None):
        cands = [c for c in candidates if c is not None]
        if not cands:
            return None
        if len(cands) == 1:
            return cands[0]
        if prefer_index is not None and candidates[prefer_index] is not None:
            preferred = candidates[prefer_index]
            # sanity check: only trust the preferred source if it's not
            # wildly inconsistent with the gridline count (catches a bad
            # OCR read even in the normally-reliable source)
            if abs((preferred - axis_min + 1) - n_grid_detected) <= max(6, 0.25 * n_grid_detected):
                return preferred
        return min(cands, key=lambda c: abs((c - axis_min + 1) - n_grid_detected))

    # Empirically, across all validated real-world samples: the anchored
    # single-label read (index 1) is the more reliable source for x_max,
    # while the whole-strip read (index 0) is more reliable for y_max.
    x_max = pick_best([x_max_strip, x_max_anchor], x_min, n_x_grid, prefer_index=1)
    y_max = pick_best([y_max_strip, y_max_anchor], y_min, n_y_grid, prefer_index=0)

    return (x_min, x_max), (y_min, y_max)


# ============================================================
# 3. GRIDLINE DETECTION + CALIBRATION REGRESSION
# ============================================================

def _detect_gridlines_local(line, window=9, min_dev=3):
    """Gridlines as LOCAL dips relative to a small rolling-median baseline,
    not one global row/column median. A global baseline collapses an
    entire shaded overlay region (e.g. 'IEC demand exceeded' band) into
    one false blob and loses the real gridlines inside it; a local
    baseline (window << gridline spacing) stays robust to that."""
    baseline = median_filter(line, size=window, mode='nearest')
    dev = baseline - line
    idx = np.where(dev > min_dev)[0]
    if len(idx) == 0:
        return np.array([])
    groups, cur = [], [idx[0]]
    for x in idx[1:]:
        if x - cur[-1] <= 2:
            cur.append(x)
        else:
            groups.append(cur); cur = [x]
    groups.append(cur)
    return np.array([np.average(g, weights=dev[g]) for g in groups])


def _robust_axis_fit(centers, n_expected, tol_px=0.4):
    centers = np.sort(centers)
    if len(centers) < 3:
        return None
    slope0 = (centers[-1] - centers[0]) / max(n_expected - 1, 1)
    intercept0 = centers[0]
    for _ in range(6):
        idx_est = (centers - intercept0) / slope0
        idx_round = np.round(idx_est)
        resid = np.abs(idx_est - idx_round)
        keep = (idx_round >= 0) & (idx_round <= n_expected - 1) & (resid < tol_px)
        if keep.sum() < 3:
            tol_px *= 1.6
            continue
        A = np.vstack([idx_round[keep], np.ones(keep.sum())]).T
        slope0, intercept0 = np.linalg.lstsq(A, centers[keep], rcond=None)[0]
    resid_final = centers - (slope0 * np.round((centers - intercept0) / slope0) + intercept0)
    return slope0, intercept0, np.abs(resid_final).mean()


def _best_fit_n_ticks(centers, n_candidates):
    """Used only when the axis max couldn't be OCR'd at all: search over
    plausible tick counts and pick whichever gives the tightest, most
    complete regression fit."""
    best = None
    for n in n_candidates:
        fit = _robust_axis_fit(centers, n)
        if fit is None:
            continue
        slope, intercept, resid = fit
        idx_est = (np.sort(centers) - intercept) / slope
        kept = ((np.round(idx_est) >= 0) & (np.round(idx_est) <= n - 1) &
                (np.abs(idx_est - np.round(idx_est)) < 0.4)).sum()
        score = (kept, -resid)
        if best is None or score > best[0]:
            best = (score, n, slope, intercept)
    return best


def _infer_tick_spacing(axis_span, n_gridlines_detected):
    """Infer the actual tick spacing (1, 2, 5, 10, ...) from the ratio of
    the axis's total span to how many gridlines were actually detected,
    snapping to the nearest 'nice' spacing value. Needed because earlier
    versions assumed every chart uses spacing=1 (a tick at every integer)
    -- true for the chart family this was built against, but NOT
    universal: a chart with a wide axis range (e.g. 0-100) may use
    spacing=5 or 10 instead, which silently broke calibration if assumed
    to be 1 (confirmed on a synthetic test chart)."""
    if n_gridlines_detected < 2 or axis_span <= 0:
        return 1
    raw_spacing = axis_span / (n_gridlines_detected - 1)
    nice_values = [1, 2, 2.5, 5, 10, 20, 25, 50, 100, 200, 250, 500, 1000]
    return min(nice_values, key=lambda v: abs(np.log(v) - np.log(raw_spacing)))


def calibrate(img, box, x_range=None, y_range=None):
    """Returns dict with x_to_col/col_to_x/y_to_row/row_to_y and fit
    quality info. x_range/y_range are (min,max) tuples; if max is None,
    falls back to a purely structural best-fit search."""
    top, bottom, left, right = box
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY).astype(float)

    x_min, x_max = x_range
    y_min, y_max = y_range

    row_probe = bottom - max(3, int((bottom - top) * 0.01))
    xg = _detect_gridlines_local(gray[row_probe, left:right + 1]) + left
    col_probe = left + max(8, int((right - left) * 0.02))
    yg = _detect_gridlines_local(gray[top:bottom + 1, col_probe]) + top

    # The plot border was observed (both chart families) to coincide
    # exactly with the axis min/max gridlines. Anchoring the fit with
    # these known endpoint pixel positions fixes edge-of-scan detection
    # misses (a gridline sitting on the very last scanned pixel can be
    # lost to boundary effects in the local-baseline filter) and also
    # makes the fit more resistant to the thick vertical v_cut-in/v_rated
    # reference lines being mistaken for real gridlines, since a spurious
    # nearby point can no longer shift where the fit anchors the ends.
    xg = np.unique(np.concatenate([xg, [left, right]]))
    yg = np.unique(np.concatenate([yg, [top, bottom]]))

    if x_max is not None:
        x_spacing = _infer_tick_spacing(x_max - x_min, len(xg))
        n_x = round((x_max - x_min) / x_spacing) + 1
        xfit = _robust_axis_fit(xg, n_x)
        x_slope, x_intercept = (xfit[0], xfit[1]) if xfit else ((right - left) / (x_max - x_min), left)
    else:
        best = _best_fit_n_ticks(xg, range(15, 60))
        n_x, x_slope, x_intercept = best[1], best[2], best[3]
        x_max = x_min + n_x - 1

    if y_max is not None:
        y_spacing = _infer_tick_spacing(y_max - y_min, len(yg))
        n_y = round((y_max - y_min) / y_spacing) + 1
        yfit = _robust_axis_fit(yg, n_y)
        y_slope_idx, y_intercept = (yfit[0], yfit[1]) if yfit else ((bottom - top) / (y_max - y_min), top)
    else:
        best = _best_fit_n_ticks(yg, range(15, 60))
        n_y, y_slope_idx, y_intercept = best[1], best[2], best[3]
        y_max = y_min + n_y - 1

    def x_to_col(x): return x_slope * (x - x_min) + x_intercept
    def col_to_x(c): return (c - x_intercept) / x_slope + x_min
    def y_to_row(y): return y_slope_idx * (y_max - y) + y_intercept
    def row_to_y(row): return y_max - (row - y_intercept) / y_slope_idx

    # Structural endpoint safeguard for dense axes. When the detected axis has
    # substantially more labelled intervals than can be stably represented by the
    # local gridline fit, endpoint drift becomes a near-constant numeric bias while
    # overlays still look correct. Use border endpoints in that condition. The
    # threshold is an algorithm stability limit, not a chart ID or pixel geometry.
    if y_max is not None and y_min is not None and (y_max - y_min) >= 35:
        y_slope_idx = (bottom - top) / float(y_max - y_min)
        y_intercept = float(top)

    return {
        'plot_box': box, 'x_range': (x_min, x_max), 'y_range': (y_min, y_max),
        'x_to_col': x_to_col, 'col_to_x': col_to_x,
        'y_to_row': y_to_row, 'row_to_y': row_to_y,
        'n_x_gridlines_found': len(xg), 'n_y_gridlines_found': len(yg),
    }


def calibration_confidence(calib, tol_frac=0.03, expected_x_max_range=None, expected_y_max_range=None):
    """Cross-check: does the fitted axis span match the detected border
    box span? If they disagree by more than tol_frac, something in the
    OCR'd range or gridline fit is wrong -- flag rather than trust.

    expected_x_max_range / expected_y_max_range are OPTIONAL plausibility
    bounds (e.g. (22,38) if you know your batch's charts always use
    x_max~30). If given, values outside them are surfaced as
    'x_max_plausible'/'y_max_plausible' = False in the returned info for
    manual review -- but this NEVER blocks `ok` on its own. A structurally
    self-consistent, correctly-detected axis range (verified against the
    image's own gridlines and border) is trusted even if it falls outside
    a prior expectation, because THAT expectation was built from a
    limited sample and a genuinely different chart in the same template
    can legitimately use a different axis range (validated: charts with
    axis max of 9, 18, 20, and 45 were all correctly detected and
    extracted, none of which would pass a band tuned to x_max~30)."""
    top, bottom, left, right = calib['plot_box']
    x_min, x_max = calib['x_range']
    y_min, y_max = calib['y_range']
    expected_w = calib['x_to_col'](x_max) - calib['x_to_col'](x_min)
    actual_w = right - left
    expected_h = abs(calib['y_to_row'](y_min) - calib['y_to_row'](y_max))
    actual_h = bottom - top
    w_err = abs(expected_w - actual_w) / max(actual_w, 1)
    h_err = abs(expected_h - actual_h) / max(actual_h, 1)
    ok = (w_err < tol_frac) and (h_err < tol_frac) and calib['n_x_gridlines_found'] >= 10 and calib['n_y_gridlines_found'] >= 10
    info = {'x_span_error_pct': round(w_err * 100, 2), 'y_span_error_pct': round(h_err * 100, 2),
            'x_range': calib['x_range'], 'y_range': calib['y_range']}
    if expected_x_max_range is not None:
        info['x_max_plausible'] = expected_x_max_range[0] <= x_max <= expected_x_max_range[1]
    if expected_y_max_range is not None:
        info['y_max_plausible'] = expected_y_max_range[0] <= y_max <= expected_y_max_range[1]
    return ok, info


# ============================================================
# 4. LEGEND COLOR AUTO-DETECTION (OCR)
# ============================================================

def _ocr_legend_lines(img):
    H, W = img.shape[:2]
    x0, x1 = int(0.80 * W), int(0.995 * W)
    y0, y1 = int(0.08 * H), int(0.30 * H)
    crop = img[y0:y1, x0:x1]
    if crop.size == 0 or not HAS_OCR:
        return [], x0
    scale = 3
    big = cv2.resize(crop, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
    gray = cv2.cvtColor(big, cv2.COLOR_BGR2GRAY)
    data = pytesseract.image_to_data(gray, output_type=pytesseract.Output.DICT)
    lines = {}
    for i, txt in enumerate(data['text']):
        if txt.strip():
            key = (data['block_num'][i], data['par_num'][i], data['line_num'][i])
            lines.setdefault(key, []).append(
                (txt, data['left'][i], data['top'][i], data['width'][i], data['height'][i]))
    out = []
    for k, v in sorted(lines.items(), key=lambda kv: min(t[2] for t in kv[1])):
        text = " ".join(t[0] for t in v)
        top = min(t[2] for t in v)
        height = max(t[2] + t[4] for t in v) - top
        left = min(t[1] for t in v)
        out.append({'text': text, 'row_c': y0 + (top + height / 2) / scale,
                     'left_px': x0 + left / scale})
    return out, x0


def _swatch_color(img, row_c, left_edge_x, x0):
    y0i, y1i = int(row_c) - 4, int(row_c) + 5
    x_start, x_end = x0, int(left_edge_x) + 15
    strip = img[y0i:y1i, x_start:x_end].astype(int)
    if strip.size == 0:
        return None
    chroma = strip.max(axis=2) - strip.min(axis=2)
    if chroma.max() < 15:
        return None
    idx = np.unravel_index(np.argmax(chroma), chroma.shape)
    return strip[idx].astype(float)


def detect_legend_colors(img):
    lines, x0 = _ocr_legend_lines(img)
    result = {'effective': None, 'ambient': None}
    for line in lines:
        low = line['text'].lower()
        if 'effective' in low and 'ambient' not in low and result['effective'] is None:
            c = _swatch_color(img, line['row_c'], line['left_px'], x0)
            if c is not None:
                result['effective'] = c
        elif 'ambient' in low and result['ambient'] is None:
            c = _swatch_color(img, line['row_c'], line['left_px'], x0)
            if c is not None:
                result['ambient'] = c

    # Sanity check: the "ambient" (pink) line's dash character is thin/
    # dotted and doesn't always get picked up as its own OCR word, which
    # can leave left_px pointing at the wrong place (e.g. the start of the
    # text itself) and sample noise instead of the real swatch -- observed
    # to produce nonsensical colors (e.g. [0,116,191], not a plausible
    # light-pink). By definition in this legend scheme, "ambient" pink is
    # a LIGHTER tint of the same red hue: its R channel should be high
    # (like red's) and its G/B channels should be HIGHER (lighter) than
    # red's, not lower/darker. Reject any detected color that violates
    # this rather than let a corrupted split-threshold silently discard
    # legitimate curve pixels downstream.
    if result['ambient'] is not None and result['effective'] is not None:
        red_g = result['effective'][1]
        pink_g = result['ambient'][1]
        pink_r = result['ambient'][2]
        if not (pink_g > red_g and pink_r > 200):
            result['ambient'] = None

    return result


FALLBACK_RED = np.array([114, 114, 255], dtype=float)
FALLBACK_PINK_G_OFFSET = 85


# ============================================================
# 5. CURVE EXTRACTION (red/pink disambiguation + continuity cleaning)
# ============================================================

def estimate_pink_from_image(img, box, red_bgr):
    """Fallback pink/ambient color estimator for when the legend swatch
    read fails (observed on real data: the ambient line's thin/dotted
    dash symbol doesn't always get picked up correctly by OCR, unlike
    the solid red and blue swatches). Scans the whole plot area for
    pixels that are reddish (R dominant) but noticeably LIGHTER than the
    known red target -- the defining trait of "ambient" being a pale
    tint of the same hue -- and takes the median as an image-specific
    estimate, which is much tighter than a fixed generic offset and was
    confirmed on real data to closely match visually-plausible pink."""
    top, bottom, left, right = box
    seg = img[top:bottom + 1, left:right + 1].reshape(-1, 3).astype(float)
    chroma = seg.max(axis=1) - seg.min(axis=1)
    red_g = red_bgr[1]
    mask = ((seg[:, 2] > 200) & (seg[:, 1] > red_g + 20) & (seg[:, 1] < 235) &
            (chroma > 20) & (np.abs(seg[:, 0] - seg[:, 1]) < 15))
    candidates = seg[mask]
    if len(candidates) < 30:
        return None
    return np.median(candidates, axis=0)


def build_red_mask_fn(red_bgr, pink_bgr=None, tol=80):
    red_g = red_bgr[1]
    split_g = (red_g + pink_bgr[1]) / 2.0 if pink_bgr is not None else red_g + FALLBACK_PINK_G_OFFSET / 2.0
    pure_marker = np.array([255, 0, 0], dtype=float)

    def weight_fn(seg):
        r, g, b = seg[:, 2], seg[:, 1], seg[:, 0]
        chroma = seg.max(axis=1) - seg.min(axis=1)

        # Hue match: R dominant, G approx B -- the defining trait of this
        # chart's red/pink color scheme, independent of exact lightness.
        gb_gap = np.abs(g - b)
        hue_ok = (r > 200) & (gb_gap < 20)

        # A pixel MORE saturated than the sampled legend swatch (lower G
        # than red_bgr's G) is still genuinely red -- likely MORE reliably
        # so, since anti-aliasing on the swatch's thin dash tends to read
        # lighter than bold curve strokes elsewhere (confirmed on real
        # data: the truest, most saturated curve pixels were being
        # excluded by a fixed-radius distance-to-swatch model). Never
        # penalize this; only taper confidence as G rises toward the
        # pink split point.
        w = np.where(
            g <= red_g,
            1.0,
            np.clip(1 - (g - red_g) / max(split_g - red_g, 1), 0, 1)
        )
        w = np.where(hue_ok, w, 0.0)
        w *= np.clip(chroma / 25, 0, 1)          # exclude grayscale gridlines
        w = np.where(g >= split_g, 0, w)          # hard cutoff into pink territory

        pure_d = np.linalg.norm(seg - pure_marker, axis=1)
        w = np.where(pure_d < 40, 0, w)           # exclude pure-blue reference lines
        return w

    return weight_fn


def _collect_column_candidates(img, calib, weight_fn):
    """Pass 1: for every column, find all plausible color-matching clusters
    (not yet choosing between them). Returns dict: col -> list of
    (run, weight_sum, mean_g, centroid_row, y_value)."""
    top, bottom, left, right = calib['plot_box']
    per_col = {}
    for col in range(left, right + 1):
        seg = img[top:bottom + 1, col, :].astype(float)
        w = weight_fn(seg)
        idx = np.where(w > 0.05)[0]
        if len(idx) == 0:
            continue
        runs, cur = [], [idx[0]]
        for i in idx[1:]:
            if i - cur[-1] <= 2:
                cur.append(i)
            else:
                runs.append(cur); cur = [i]
        runs.append(cur)

        run_stats = []
        for run in runs:
            weight_sum = w[run].sum()
            if max(w[run]) < 0.12:
                continue
            mean_g = np.mean(seg[run, 1])
            rows_local = np.array(run)
            centroid = np.average(rows_local, weights=w[run]) + top
            y_val = calib['row_to_y'](centroid)
            run_stats.append((run, weight_sum, mean_g, centroid, y_val))
        if run_stats:
            per_col[col] = run_stats
    return per_col


def extract_raw_curve(img, calib, weight_fn):
    """Two-pass extraction with genuine LOOK-AHEAD, not just look-behind.

    A purely left-to-right greedy tracker (tried first) has a real blind
    spot: if a wrong candidate happens to win several consecutive columns
    in a row (observed on real data -- a dotted 'ambient' line can shadow
    the true curve for a run of columns, not just isolated ones), the
    tracker's own trend estimate gets pulled toward the wrong track and
    then keeps "correctly" following it. It has no way to know a better
    answer resumes a few columns later.

    Fix: first pass picks the single most-saturated cluster per column
    (ignoring ambiguity) to build a ROUGH reference curve. A robust
    rolling-MEDIAN smooth of that rough curve is largely immune to a
    short wrong run (a median naturally ignores a minority pattern
    within its window, whether that minority is "before" or "after" the
    point in question -- true look-ahead AND look-behind). The second
    pass then re-resolves every column's real candidates against this
    smoothed reference instead of a single neighboring point, which
    correctly recovers columns a forward-only tracker would get stuck on.
    """
    candidates_by_col = _collect_column_candidates(img, calib, weight_fn)
    if not candidates_by_col:
        return {}

    cols_sorted = sorted(candidates_by_col)

    # --- Pass 1: rough curve via per-column saturation heuristic ---
    rough_y = {}
    for col in cols_sorted:
        stats = candidates_by_col[col]
        max_weight = max(r[1] for r in stats)
        cands = [r for r in stats if r[1] >= 0.35 * max_weight]
        chosen = min(cands, key=lambda r: r[2])  # most saturated
        rough_y[col] = chosen[4]

    # --- Robust rolling-median reference curve (true look-ahead/behind) ---
    cols_arr = np.array(cols_sorted)
    y_arr = np.array([rough_y[c] for c in cols_sorted])
    ref_y = np.empty_like(y_arr)
    HALF_WIN_PX = 12  # in source columns, not raw-candidate-list index
    for i, c in enumerate(cols_arr):
        m = (cols_arr >= c - HALF_WIN_PX) & (cols_arr <= c + HALF_WIN_PX)
        ref_y[i] = np.median(y_arr[m])
    ref_map = dict(zip(cols_sorted, ref_y))

    # --- Pass 2: re-resolve each column's real candidates against the
    #     robust reference instead of a single neighbor ---
    raw = {}
    MAX_DEV = 1.5  # data units; reject a column entirely if nothing plausible is close to the reference
    for col in cols_sorted:
        stats = candidates_by_col[col]
        ref = ref_map[col]

        # Candidate-strength guard: line/grid intersections can create a weak
        # reddish one-pixel candidate closer to the rolling reference than the
        # real curve. Keep only candidates carrying at least 35% of the strongest
        # colour evidence in that column before continuity selection. This rule is
        # relative (not tied to any graph, axis, resolution, x-value or colour),
        # so it remains fully dynamic across chart families.
        max_weight = max(r[1] for r in stats)
        strong_stats = [r for r in stats if r[1] >= 0.35 * max_weight]
        chosen = min(strong_stats, key=lambda r: abs(r[4] - ref))
        if abs(chosen[4] - ref) > MAX_DEV:
            continue
        raw[col] = chosen[3]

    return raw


def clean_curve(raw_cols, calib, window=9, max_dev_data_units=0.45, edge_protect=3):
    """Continuity-based outlier rejection, EXCEPT at the first/last
    `edge_protect` detected columns, which are always kept regardless of
    local deviation. Rationale (found on real data): a curve can have a
    genuinely steep initial/final slope (e.g. dropping several TI% within
    just 2-3 pixel-columns right at its start point) that looks identical
    to contamination under a naive local-neighborhood check, and was
    being wrongly discarded -- silently replacing the correct boundary
    value with a biased-toward-the-interior one. The crossing-line
    contamination this filter is designed to catch happens mid-curve,
    not at the very first/last few detected columns, so exempting the
    edges from rejection is safe."""
    if not raw_cols:
        return {}
    cols = np.array(sorted(raw_cols))
    rows = np.array([raw_cols[c] for c in cols])
    ys = np.array([calib['row_to_y'](r) for r in rows])
    half = window // 2
    keep = np.ones(len(cols), dtype=bool)
    n = len(cols)
    for i in range(n):
        if i < edge_protect or i >= n - edge_protect:
            continue  # always keep boundary points
        lo, hi = max(0, i - half), min(n, i + half + 1)
        nb = np.delete(ys[lo:hi], i - lo)
        if len(nb) == 0:
            continue
        if abs(ys[i] - np.median(nb)) > max_dev_data_units:
            keep[i] = False
    cols_c, ys_c = cols[keep], ys[keep]
    if len(cols_c) < 2:
        return raw_cols
    full_cols = np.arange(cols_c[0], cols_c[-1] + 1)
    ys_interp = np.interp(full_cols, cols_c, ys_c)
    return {c: y for c, y in zip(full_cols, ys_interp)}


def sample_at_x(cleaned_col_to_y, calib, x_query, half_window_px=1.5):
    if not cleaned_col_to_y:
        return None
    c_center = calib['x_to_col'](x_query)
    c0, c1 = int(np.floor(c_center - half_window_px)), int(np.ceil(c_center + half_window_px))
    vals, wts = [], []
    for c in range(c0, c1 + 1):
        if c in cleaned_col_to_y:
            wcol = max(0.01, 1 - abs(c - c_center) / (half_window_px + 1))
            vals.append(cleaned_col_to_y[c]); wts.append(wcol)
    if vals:
        return float(np.average(vals, weights=wts))

    # Requested x falls outside the curve's detected column range (common
    # right at the curve's start/end point, where the visible line's tip
    # is a few px further in than the mathematical x-axis position due to
    # line width / anti-aliasing). Rather than widening the search window
    # and silently grabbing a DISTANT, unrelated point on the curve --
    # which was observed to produce a confidently wrong answer, worse
    # than returning nothing -- extrapolate linearly from the nearest 2-3
    # genuinely-detected points, and only within a small distance (curves
    # in this chart family are steep near the edges, so extrapolating far
    # is unsafe).
    cols_sorted = np.array(sorted(cleaned_col_to_y))
    MAX_EXTRAPOLATE_PX = 20
    if c_center < cols_sorted[0] and cols_sorted[0] - c_center <= MAX_EXTRAPOLATE_PX:
        near = cols_sorted[:3]
        ys = np.array([cleaned_col_to_y[c] for c in near])
        slope = np.polyfit(near, ys, 1)[0] if len(near) >= 2 else 0
        return float(cleaned_col_to_y[cols_sorted[0]] + slope * (c_center - cols_sorted[0]))
    if c_center > cols_sorted[-1] and c_center - cols_sorted[-1] <= MAX_EXTRAPOLATE_PX:
        near = cols_sorted[-3:]
        ys = np.array([cleaned_col_to_y[c] for c in near])
        slope = np.polyfit(near, ys, 1)[0] if len(near) >= 2 else 0
        return float(cleaned_col_to_y[cols_sorted[-1]] + slope * (c_center - cols_sorted[-1]))

    return None


# ============================================================
# 6. TITLE DETECTION (with filename cross-validation)
# ============================================================

def detect_title(img):
    if not HAS_OCR:
        return None
    H, W = img.shape[:2]
    crop = img[0:int(0.08 * H), int(0.25 * W):int(0.75 * W)]
    gray0 = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
    for fx in (4, 3, 5, 6, 2):
        big = cv2.resize(gray0, None, fx=fx, fy=fx, interpolation=cv2.INTER_CUBIC)
        for cfg in ('', '--psm 7', '--psm 6'):
            txt = pytesseract.image_to_string(big, config=cfg).strip().replace('\n', ' ').strip()
            if txt and len(txt) <= 40:
                return txt
    return None


def resolve_graph_name(img, filename_stem):
    ocr_name = detect_title(img)
    if not ocr_name:
        return filename_stem, ocr_name, "filename (OCR found nothing)", False
    norm_ocr = ocr_name.replace(' ', '').upper()
    norm_file = filename_stem.replace(' ', '').upper()
    agrees = norm_file in norm_ocr or norm_ocr in norm_file or norm_ocr == norm_file
    if agrees:
        return ocr_name, ocr_name, "OCR (matches filename)", False
    return filename_stem, ocr_name, "filename (OCR disagreed)", True


# ============================================================
# 7. TOP-LEVEL: process one image
# ============================================================

def process_image(path, target_x, debug=False):
    img = cv2.imread(path)
    if img is None:
        raise ValueError(f"Could not read image: {path}")

    img, norm_reason = normalize_image(img)
    if img is None:
        raise ValueError(f"Image too low-resolution to process reliably: {norm_reason}")

    box = detect_plot_box(cv2.cvtColor(img, cv2.COLOR_BGR2GRAY))
    x_range, y_range = detect_axis_range(img, box)
    calib = calibrate(img, box, x_range, y_range)
    conf_ok, conf_info = calibration_confidence(calib)

    colors = detect_legend_colors(img)
    red_bgr = colors['effective'] if colors['effective'] is not None else FALLBACK_RED
    pink_bgr = colors['ambient']
    if pink_bgr is None:
        pink_bgr = estimate_pink_from_image(img, box, red_bgr)

    weight_fn = build_red_mask_fn(red_bgr, pink_bgr)
    raw = extract_raw_curve(img, calib, weight_fn)
    cleaned = clean_curve(raw, calib)

    results = {}
    for x in target_x:
        results[x] = sample_at_x(cleaned, calib, x) if conf_ok else None

    if debug:
        return results, {'red_bgr': red_bgr, 'pink_bgr': pink_bgr, 'calib': calib,
                          'calibration_ok': conf_ok, 'calibration_info': conf_info,
                          'raw_n': len(raw), 'cleaned_n': len(cleaned)}
    return results, conf_ok, conf_info


# ============================================================
# 8. BATCH RUNNER
# ============================================================

def find_image_files(folder='.', patterns=('*.png', '*.jpg', '*.jpeg', '*.PNG', '*.JPG', '*.JPEG')):
    files = []
    for pat in patterns:
        files.extend(glob.glob(os.path.join(folder, pat)))
    return sorted(set(files))


def process_folder(folder, target_x, out_xlsx, patterns=('*.png', '*.jpg', '*.jpeg', '*.PNG', '*.JPG', '*.JPEG')):
    files = find_image_files(folder, patterns)
    df = pd.DataFrame({"Wind speed (m/s)": target_x})
    audit_rows = []

    for f in files:
        stem = os.path.splitext(os.path.basename(f))[0]
        print(f"--- {os.path.basename(f)} ---")
        img = cv2.imread(f)
        if img is None:
            print("  FAILED: unreadable image file")
            df[stem] = [np.nan] * len(target_x)
            audit_rows.append({"file": stem, "status": "FAILED: unreadable image file",
                                "calibration_ok": False})
            continue

        img, norm_reason = normalize_image(img)
        if img is None:
            print(f"  FAILED: {norm_reason}")
            df[stem] = [np.nan] * len(target_x)
            audit_rows.append({"file": stem, "status": f"FAILED: {norm_reason}",
                                "calibration_ok": False})
            continue

        try:
            box = detect_plot_box(cv2.cvtColor(img, cv2.COLOR_BGR2GRAY))
            x_range, y_range = detect_axis_range(img, box)
            calib = calibrate(img, box, x_range, y_range)
            conf_ok, conf_info = calibration_confidence(calib)

            if not conf_ok:
                print(f"  WARNING: calibration low-confidence -- {conf_info}")
                df[stem] = [np.nan] * len(target_x)
                audit_rows.append({"file": stem, "status": "FAILED: calibration low-confidence",
                                    "calibration_ok": False, **conf_info})
                continue

            colors = detect_legend_colors(img)
            red_bgr = colors['effective'] if colors['effective'] is not None else FALLBACK_RED
            pink_bgr = colors['ambient']
            if pink_bgr is None:
                pink_bgr = estimate_pink_from_image(img, box, red_bgr)
            weight_fn = build_red_mask_fn(red_bgr, pink_bgr)
            raw = extract_raw_curve(img, calib, weight_fn)
            cleaned = clean_curve(raw, calib)

            vals = [sample_at_x(cleaned, calib, x) for x in target_x]
            df[stem] = [round(v, 3) if v is not None else np.nan for v in vals]

            graph_name, ocr_raw, name_source, mismatch = resolve_graph_name(img, stem)
            n_unresolved = sum(1 for v in vals if v is None)
            status = "OK" if n_unresolved == 0 else f"OK with {n_unresolved} unresolved point(s)"
            print(f"  graph_name={graph_name} [{name_source}]  x_range={conf_info['x_range']} "
                  f"y_range={conf_info['y_range']}  status={status}")
            audit_rows.append({"file": stem, "graph_name": graph_name, "status": status,
                                "calibration_ok": True, "unresolved_points": n_unresolved,
                                **conf_info})
        except Exception as e:
            print(f"  FAILED: {type(e).__name__}: {e}")
            df[stem] = [np.nan] * len(target_x)
            audit_rows.append({"file": stem, "status": f"FAILED: {type(e).__name__}: {e}",
                                "calibration_ok": False})

    df_audit = pd.DataFrame(audit_rows)
    with pd.ExcelWriter(out_xlsx, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Effective_TI', index=False)
        df_audit.to_excel(writer, sheet_name='QA_Audit_Log', index=False)
    print(f"\nSaved {out_xlsx}")
    return df, df_audit


def _check_tesseract_ready():
    """Fail fast with a clear, actionable message if OCR isn't actually
    reachable, rather than letting every image silently degrade to
    filename-only fallbacks with no explanation."""
    if not HAS_OCR:
        print("WARNING: pytesseract is not installed (pip install pytesseract). "
              "OCR-based axis-range detection and legend-color detection will be "
              "SKIPPED for all images -- results will rely entirely on structural "
              "gridline-counting fallbacks, which is less reliable. Install "
              "pytesseract and the Tesseract OCR engine for full accuracy.")
        return
    try:
        import pytesseract
        pytesseract.get_tesseract_version()
        print(f"Tesseract OK -- using: {pytesseract.pytesseract.tesseract_cmd}")
    except Exception as e:
        print(f"WARNING: pytesseract is installed but could not reach the Tesseract "
              f"engine ({e}). Set TESSERACT_CMD at the top of this script to the full "
              f"path of your tesseract executable. Proceeding with reduced accuracy "
              f"(structural fallbacks only).")



# ============================================================
# 9. V2 ENHANCEMENTS: direct Excel image reading, structural axis
#    arbitration, robust title ID, overlays and accuracy workbook
# ============================================================
import io
import re
import difflib
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

INPUT_XLSX = "all_output.xlsx"
IMAGE_SHEET = "all_images"
V2_OUTPUT_XLSX = "effective_ti_output_v7.xlsx"
OVERLAY_DIR = "processed_overlays_v5"

# Generic title-token grammar. No plant/site/project prefixes are listed here.
# A graph ID must contain both a letter and a digit and may contain separators.
# Examples accepted automatically: ABC-123, WTG_07A, MP540(S), MP8&12,
# SITE-A-UNIT-004, X1, 123-AB, etc.
_GENERIC_ID_TOKEN = re.compile(
    r'(?<![A-Z0-9])(?=[A-Z0-9_&()\-/ ]{2,60}(?![A-Z0-9]))'
    r'(?=[A-Z0-9_&()\-/ ]*[A-Z])(?=[A-Z0-9_&()\-/ ]*\d)'
    r'[A-Z0-9]+(?:[-_/ &()][A-Z0-9()&]+){0,6}(?![A-Z0-9])'
)

# These are chart vocabulary, not ID families. They are used only to reject
# obvious axis/legend/subtitle text from title candidates.
_TITLE_NOISE = {
    'IEC','CLASS','MAST','EFFECTIVE','AMBIENT','DEMAND','EXCEEDED','WIND',
    'SPEED','TURBULENCE','PERCENT','MEAN','NORMAL','MODEL','HEIGHT','HUB',
    'NTM','ETM','TI','M','MS'
}

def _decode_excel_image(xl_image):
    """Decode the original embedded image bytes; no re-render or PNG save."""
    arr = np.frombuffer(xl_image._data(), dtype=np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    if img is None:
        raise ValueError("Could not decode embedded Excel image")
    return img


def _text_variants(gray):
    out = [gray]
    out.append(cv2.normalize(gray, None, 0, 255, cv2.NORM_MINMAX))
    out.append(cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1])
    out.append(cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                     cv2.THRESH_BINARY, 31, 9))
    return out


def _normalise_id(text):
    text = str(text).upper().strip()
    text = re.sub(r'[^A-Z0-9_&()\-/ ]+', '', text)
    text = re.sub(r'\s+', '-', text)
    text = re.sub(r'[_/]+', '-', text)
    text = re.sub(r'-{2,}', '-', text)
    return text.strip('-_ /')


def _id_candidate_score(value, confidence, y_center, image_h, candidate_names=None):
    """Score an OCR token structurally; no known ID prefix is required."""
    value = _normalise_id(value)
    if (not value or not re.search(r'[A-Z]', value) or not re.search(r'\d', value)
            or y_center > image_h * 0.13):
        return None
    words = set(re.findall(r'[A-Z]+', value))
    if words & _TITLE_NOISE:
        return None
    # Prefer compact top-most tokens with good OCR confidence. Excessively long
    # strings are more likely to be a subtitle sentence than a graph ID.
    compactness = 1.0 / (1.0 + max(0, len(value) - 20) * 0.08)
    top_prior = max(0.0, 1.0 - y_center / max(image_h * 0.22, 1))
    score = max(float(confidence), 0.0) / 100.0 + 0.55 * top_prior + 0.25 * compactness
    resolved = value
    if candidate_names:
        norm = lambda z: re.sub(r'[^A-Z0-9]', '', str(z).upper())
        sims = [(difflib.SequenceMatcher(None, norm(value), norm(c)).ratio(), str(c))
                for c in candidate_names]
        sim, candidate = max(sims, default=(0.0, value))
        # Correct only near-exact OCR differences. A loose match can silently map
        # a new production ID to an unrelated historical ID.
        if sim >= 0.88:
            resolved = candidate
            score += 0.35 * sim
    return score, resolved


def _ocr_title_pass(gray0, image_h, candidate_names, scale, psm, use_threshold=False):
    """Run one focused OCR pass and return scored generic ID candidates."""
    src = gray0
    if use_threshold:
        src = cv2.threshold(gray0, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
    big = cv2.resize(src, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
    data = pytesseract.image_to_data(
        big, config=f'--psm {psm}', output_type=pytesseract.Output.DICT)
    votes, lines = [], {}
    for i, raw in enumerate(data['text']):
        raw = raw.strip().upper()
        if not raw:
            continue
        key = (data['block_num'][i], data['par_num'][i], data['line_num'][i])
        lines.setdefault(key, []).append(i)
        value = re.sub(r'[^A-Z0-9_&()\-/]', '', raw)
        sc = _id_candidate_score(
            value, data['conf'][i],
            (data['top'][i] + data['height'][i] / 2) / scale,
            image_h, candidate_names)
        if sc:
            votes.append(sc)
    # Join short neighbouring words when the ID was rendered with spaces.
    for idxs in lines.values():
        idxs = sorted(idxs, key=lambda i: data['left'][i])
        for width in (2, 3):
            for k in range(0, len(idxs) - width + 1):
                take = idxs[k:k+width]
                gaps = [data['left'][take[j+1]] -
                        (data['left'][take[j]] + data['width'][take[j]])
                        for j in range(len(take)-1)]
                if any(g > 2.5 * max(data['height'][i] for i in take) for g in gaps):
                    continue
                value = '-'.join(data['text'][i].strip().upper() for i in take)
                conf = np.mean([max(float(data['conf'][i]), 0) for i in take])
                yc = np.mean([(data['top'][i] + data['height'][i]/2)/scale for i in take])
                sc = _id_candidate_score(value, conf, yc, image_h, candidate_names)
                if sc:
                    votes.append(sc)
    return votes


def _normalise_title_id(text):
    """Conservative OCR cleanup with no graph-family assumptions."""
    text = str(text).upper().strip()
    for old in ('\u2010','\u2011','\u2012','\u2013','\u2014','_','/'):
        text = text.replace(old, '-')
    text = re.sub(r'[^A-Z0-9&()\- ]+', '', text)
    text = re.sub(r'\s+', '-', text)
    return re.sub(r'-{2,}', '-', text).strip('-')


def _valid_title_id(value):
    if not value or len(value) < 2 or len(value) > 45:
        return False
    if not re.search(r'[A-Z]', value) or not re.search(r'\d', value):
        return False
    noise = {'MAST','IEC','CLASS','EFFECTIVE','AMBIENT','DEMAND','WIND',
             'SPEED','TURBULENCE','PERCENT','MEAN','MODEL','HEIGHT','HUB',
             'NTM','ETM','TI'}
    return not bool(set(re.findall(r'[A-Z]+', value)) & noise)


def detect_title(img, candidate_names=None, return_details=False):
    """Advanced generic title OCR using a focused title band and threshold voting.

    Reading only the extreme top band prevents subtitle/mast text from competing
    with the graph ID. Three nearby binarisation levels compensate for different
    antialiasing/rendering without assuming any ID prefix or numeric sequence.
    """
    if not HAS_OCR:
        result = (None, {'confidence': 0, 'votes': 0, 'candidates': ''})
        return result if return_details else None
    H, W = img.shape[:2]
    band_h = max(38, min(H, int(round(0.078 * H))))
    gray = cv2.cvtColor(img[:band_h, :], cv2.COLOR_BGR2GRAY)
    votes = []
    # These are image binarisation levels, not graph/data assumptions. Majority
    # voting makes the read robust to anti-aliasing and thin glyph strokes.
    for threshold in (200, 215, 220):
        binary = cv2.threshold(gray, threshold, 255, cv2.THRESH_BINARY)[1]
        big = cv2.resize(binary, None, fx=5.0, fy=5.0,
                         interpolation=cv2.INTER_CUBIC)
        text = pytesseract.image_to_string(big, config='--psm 11').upper()
        for line in text.splitlines():
            value = _normalise_title_id(line)
            if _valid_title_id(value):
                votes.append(value)
                break
    if not votes:
        result = (None, {'confidence': 0, 'votes': 0, 'candidates': ''})
        return result if return_details else None

    groups = {}
    for value in votes:
        key = re.sub(r'[^A-Z0-9]', '', value)
        groups.setdefault(key, []).append(value)
    ranked = sorted(groups.values(), key=lambda items: (len(items), len(max(items,key=len))), reverse=True)
    winners = ranked[0]
    title = max(winners, key=len)

    if candidate_names:
        norm = lambda z: re.sub(r'[^A-Z0-9]', '', str(z).upper())
        sims = [(difflib.SequenceMatcher(None, norm(title), norm(c)).ratio(), str(c))
                for c in candidate_names]
        sim, candidate = max(sims, default=(0.0, title))
        if sim >= 0.92:
            title = candidate
    agreement = len(winners)
    qa = {'confidence': round(100.0 * agreement / len(votes), 1),
          'votes': agreement, 'candidates': ' | '.join(votes)}
    return (title, qa) if return_details else title


def _grid_centers_consensus(img, box, axis, probes=13):
    """Detect grid positions at many clean probes then cluster by pixel location.
    Border lines are force-included because the user confirmed they are identical.
    """
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY).astype(float)
    top, bottom, left, right = box
    all_pos = []
    if axis == 'x':
        for frac in np.linspace(.18, .92, probes):
            r = int(round(top + frac*(bottom-top)))
            all_pos.extend((_detect_gridlines_local(gray[r, left:right+1]) + left).tolist())
        lo, hi = left, right
    else:
        for frac in np.linspace(.08, .82, probes):
            c = int(round(left + frac*(right-left)))
            all_pos.extend((_detect_gridlines_local(gray[top:bottom+1, c]) + top).tolist())
        lo, hi = top, bottom
    all_pos.extend([lo, hi])
    vals = sorted(all_pos)
    if not vals:
        return np.array([lo, hi], float)
    groups = [[vals[0]]]
    for v in vals[1:]:
        if v - np.mean(groups[-1]) <= 2.5:
            groups[-1].append(v)
        else:
            groups.append([v])
    # retain recurrent lines; endpoints always retained
    centers = [float(np.median(g)) for g in groups
               if len(g) >= max(2, probes//4) or abs(np.median(g)-lo)<3 or abs(np.median(g)-hi)<3]
    centers = np.array(sorted(set(round(v,2) for v in centers)), float)
    # Remove non-periodic clutter with spacing-consistency voting.
    if len(centers) >= 5:
        diffs = np.diff(centers)
        small = diffs[(diffs > 3) & (diffs < (hi-lo)/2)]
        if len(small):
            # mode-like estimate from pairwise candidate spacings
            hist, edges = np.histogram(small, bins=max(8, min(60, int(hi-lo)//4)))
            step = (edges[np.argmax(hist)] + edges[np.argmax(hist)+1])/2
            # score offsets and rebuild the full regular lattice anchored to borders
            candidates = [step, (hi-lo)/round((hi-lo)/step)] if step>0 else []
            best = None
            for s in candidates:
                n = max(2, int(round((hi-lo)/s))+1)
                lattice = np.linspace(lo, hi, n)
                err = np.mean([np.min(abs(centers-v)) for v in lattice])
                score = (err, -n)
                if best is None or score < best[0]: best=(score,lattice)
            if best is not None and best[0][0] <= 3.0:
                centers = best[1]
    return centers


_original_detect_axis_range = detect_axis_range

def detect_axis_range(img, box, default_x_min=0, default_y_min=0):
    """OCR-first, structural cross-check second. Existing good OCR decisions are
    preserved. If OCR is absent or inconsistent, border-to-border grid-box counts
    provide a deterministic fallback. This changes calibration only when needed.
    """
    xr, yr = _original_detect_axis_range(img, box, default_x_min, default_y_min)
    xg = _grid_centers_consensus(img, box, 'x')
    yg = _grid_centers_consensus(img, box, 'y')
    def arbitrate(rng, centers, axis):
        amin, amax = rng
        boxes = max(1, len(centers)-1)
        if amax is None or amax <= amin:
            return (amin, amin + boxes)
        span = amax-amin
        nice = _infer_tick_spacing(span, len(centers))
        implied_boxes = span/nice if nice else span
        mismatch = abs(implied_boxes-boxes)/max(boxes,1)
        # Keep existing result unless structural evidence is clearly contradictory.
        if mismatch <= 0.14:
            return rng
        # Unit-grid fallback is appropriate only when line density itself supports it.
        if boxes >= 8 and boxes <= 80:
            structural_max = amin + boxes
            if amax > amin and abs(structural_max-amax) <= max(3, .12*amax):
                return (amin, structural_max)
        return rng
    return arbitrate(xr,xg,'x'), arbitrate(yr,yg,'y')


def process_image_array(img, target_x):
    """Same extraction pipeline as V1, but operates directly on decoded bytes."""
    img, reason = normalize_image(img)
    if img is None: raise ValueError(reason)
    box = detect_plot_box(cv2.cvtColor(img, cv2.COLOR_BGR2GRAY))
    xr, yr = detect_axis_range(img, box)
    calib = calibrate(img, box, xr, yr)
    ok, info = calibration_confidence(calib)
    colors = detect_legend_colors(img)
    red = colors['effective'] if colors['effective'] is not None else FALLBACK_RED
    pink = colors['ambient']
    if pink is None: pink = estimate_pink_from_image(img, box, red)
    raw = extract_raw_curve(img, calib, build_red_mask_fn(red, pink))
    cleaned = clean_curve(raw, calib)
    vals = {x: sample_at_x(cleaned, calib, x) if ok else None for x in target_x}
    return vals, {'image':img,'box':box,'calib':calib,'ok':ok,'info':info,
                  'red_bgr':red,'pink_bgr':pink,'raw_n':len(raw),'cleaned_n':len(cleaned)}


def build_overlay_image(meta, values, include_labels=True):
    """Return an overlay as a BGR NumPy image for a future GUI.

    Nothing is written to disk or embedded in Excel. A GUI can display the return
    value directly with OpenCV, Tkinter/PIL, PyQt, Streamlit, etc.
    """
    out = meta['image'].copy()
    cal = meta['calib']
    top, bottom, left, right = meta['box']
    cv2.rectangle(out, (left, top), (right, bottom), (0, 180, 0), 2)
    for x, y in values.items():
        if y is None:
            continue
        c = int(round(cal['x_to_col'](x)))
        r = int(round(cal['y_to_row'](y)))
        cv2.drawMarker(out, (c, r), (255, 0, 255),
                       cv2.MARKER_TILTED_CROSS, 16, 2, cv2.LINE_AA)
        if include_labels:
            cv2.putText(out, str(x), (c+5, r-5), cv2.FONT_HERSHEY_SIMPLEX,
                        .35, (40, 40, 40), 1, cv2.LINE_AA)
    return out


def _ground_truth_tables(wb):
    gt={}
    candidates=[]
    for ws in wb.worksheets:
        if ws.title.lower() in ('all_images','effective_ti','qa_audit_log','accuracy_check','processed_overlays'):
            continue
        rows=list(ws.iter_rows(values_only=True))
        if not rows: continue
        # row-oriented (MAG): first row contains x values; first col contains IDs
        if len(rows[0])>2 and isinstance(rows[0][1],(int,float)):
            xs=list(rows[0][1:])
            for row in rows[1:]:
                if not row or row[0] is None: continue
                name=str(row[0]); candidates.append(name)
                gt[name]={float(x):v for x,v in zip(xs,row[1:]) if x is not None and v is not None}
        else: # column-oriented
            headers=list(rows[0])
            for j,name in enumerate(headers[1:],1):
                if name is None: continue
                name=str(name); candidates.append(name); d={}
                for row in rows[1:]:
                    if row[0] is not None and j<len(row) and row[j] is not None:
                        d[float(row[0])]=row[j]
                gt[name]=d
    return gt,candidates


def _style_workbook(path):
    wb=load_workbook(path)
    navy='17365D'; blue='D9EAF7'; green='E2F0D9'; orange='FCE4D6'
    for ws in wb.worksheets:
        ws.freeze_panes='A2'; ws.sheet_view.showGridLines=False
        for cell in ws[1]:
            cell.fill=PatternFill('solid',fgColor=navy); cell.font=Font(color='FFFFFF',bold=True)
            cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        ws.auto_filter.ref=ws.dimensions
        for col in range(1,ws.max_column+1):
            maxlen=max((len(str(ws.cell(r,col).value or '')) for r in range(1,min(ws.max_row,200)+1)),default=8)
            ws.column_dimensions[get_column_letter(col)].width=min(max(maxlen+2,11),32)
    if 'Accuracy_Check' in wb.sheetnames:
        ws=wb['Accuracy_Check']
        # AbsErr column F
        ws.conditional_formatting.add(f'F2:F{ws.max_row}',ColorScaleRule(start_type='min',start_color='E2F0D9',mid_type='percentile',mid_value=70,mid_color='FFF2CC',end_type='max',end_color='F4CCCC'))
        for c in ws['C'][1:]: c.number_format='0.###'
        for c in ws['D'][1:]: c.number_format='0.###'
        for c in ws['F'][1:]: c.number_format='0.000'
    wb.save(path)


def process_excel_workbook(input_xlsx, output_xlsx, image_sheet=IMAGE_SHEET,
                           target_x=TARGET_X, overlay_dir=None):
    """Production batch: Excel embedded images -> Effective_TI + QA_Audit_Log.

    Deliberately excludes ground-truth analysis and overlay generation. The
    overlay_dir argument is retained only for backward call compatibility and is
    ignored. This keeps the numeric extraction pipeline identical while removing
    non-production I/O and repeated validation work.
    """
    wb = load_workbook(input_xlsx, data_only=False)
    if image_sheet not in wb.sheetnames:
        raise ValueError(f"Missing sheet: {image_sheet}")
    images = wb[image_sheet]._images
    records, audit = [], []
    seen_titles = set()

    for idx, xlimg in enumerate(images, 1):
        started = __import__('time').perf_counter()
        try:
            img = _decode_excel_image(xlimg)
            # Title OCR and numeric extraction are independent. Candidate names are
            # intentionally omitted so new production IDs are never forced to an
            # old workbook dictionary.
            title, title_qa = detect_title(img, candidate_names=None, return_details=True)
            vals, meta = process_image_array(img, target_x)
            graph = title or f'Image_{idx:03d}'
            duplicate = graph in seen_titles
            seen_titles.add(graph)

            rec = {'Graph': graph}
            rec.update({str(x): round(vals[x], 3) if vals[x] is not None else None
                        for x in target_x})
            records.append(rec)

            unresolved = sum(v is None for v in vals.values())
            if not meta['ok']:
                status = 'LOW CONFIDENCE'
            elif duplicate:
                status = 'REVIEW: duplicate OCR title'
            elif unresolved:
                status = f'OK with {unresolved} unresolved point(s)'
            else:
                status = 'OK'
            audit.append({
                'Image_No': idx, 'Graph': graph, 'Status': status,
                'Title_Duplicate': duplicate,
                'Title_OCR_Confidence': title_qa.get('confidence'),
                'Title_OCR_Votes': title_qa.get('votes'),
                'X_Range': str(meta['calib']['x_range']),
                'Y_Range': str(meta['calib']['y_range']),
                'X_Grid_Lines': meta['calib']['n_x_gridlines_found'],
                'Y_Grid_Lines': meta['calib']['n_y_gridlines_found'],
                'Raw_Points': meta['raw_n'], 'Cleaned_Points': meta['cleaned_n'],
                'Unresolved_Points': unresolved,
                'Processing_Seconds': round(__import__('time').perf_counter()-started, 2)
            })
        except Exception as e:
            audit.append({
                'Image_No': idx, 'Graph': f'Image_{idx:03d}',
                'Status': f'FAILED: {type(e).__name__}: {e}',
                'Title_Duplicate': False,
                'Processing_Seconds': round(__import__('time').perf_counter()-started, 2)
            })

    with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
        pd.DataFrame(records).to_excel(writer, sheet_name='Effective_TI', index=False)
        pd.DataFrame(audit).to_excel(writer, sheet_name='QA_Audit_Log', index=False)
    _style_workbook(output_xlsx)
    return pd.DataFrame(records), pd.DataFrame(audit)



if __name__ == '__main__':
    _check_tesseract_ready()
    input_path = INPUT_XLSX if os.path.isabs(INPUT_XLSX) else os.path.join(INPUT_FOLDER, INPUT_XLSX)
    output_path = V2_OUTPUT_XLSX if os.path.isabs(V2_OUTPUT_XLSX) else os.path.join(INPUT_FOLDER, V2_OUTPUT_XLSX)
    process_excel_workbook(input_path, output_path, IMAGE_SHEET, TARGET_X)
    print(f"Saved {output_path}")
